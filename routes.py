import os
from flask import render_template, request, redirect, url_for, flash, jsonify, session, send_file
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash
from functools import wraps
from app import app, db
from models import ReflixTracking, LogFile, MonitoredFolder, MonitoredFileState, MonitorInstance
from log_parser import ReflixLogParser
from folder_monitor import get_monitor, start_monitor, stop_monitor, is_monitor_running
import logging

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {'txt', 'log'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Authentication helpers
def get_admin_password():
    """Get admin password from environment variable."""
    return os.environ.get('ADMIN_PASSWORD', 'admin123')  # Default for dev only

def is_authenticated():
    """Check if user is authenticated as admin."""
    return session.get('admin_authenticated', False)

def require_admin_auth(f):
    """Decorator to require admin authentication for routes."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_authenticated():
            flash('Admin authentication required to access this page', 'warning')
            return redirect(url_for('admin_login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

def is_safe_path(path):
    """Validate that the path is safe to monitor."""
    # Convert to absolute path
    abs_path = os.path.abspath(path)
    
    # First check allowed prefixes (/tmp or current working directory)
    cwd = os.getcwd()
    allowed_prefixes = ['/tmp', cwd]
    
    for allowed in allowed_prefixes:
        if abs_path.startswith(os.path.abspath(allowed)):
            return True
    
    # List of forbidden paths/prefixes (only reject if not in allowed prefixes)
    forbidden_paths = [
        '/',
        '/root',
        '/home',
        '/etc',
        '/usr',
        '/var',
        '/bin',
        '/sbin',
        '/boot',
        '/proc',
        '/sys',
        '/dev',
    ]
    
    # Check if path starts with any forbidden prefix
    for forbidden in forbidden_paths:
        if abs_path.startswith(forbidden):
            return False
    
    return False


def validate_path_access(path, access_mode):
    """Validate folder path based on access mode."""
    import platform
    
    try:
        # Resolve the path to get the absolute path
        resolved_path = os.path.realpath(path)
        
        if access_mode == 'safe':
            # Use existing safe path logic
            return is_safe_path(path)
        
        elif access_mode == 'home_desktop':
            # Allow home directory and Desktop folder
            home_dir = os.path.expanduser('~')
            home_realpath = os.path.realpath(home_dir)
            
            # Desktop folder location depends on OS
            if platform.system() == 'Windows':
                desktop_path = os.path.join(home_dir, 'Desktop')
            elif platform.system() == 'Darwin':  # macOS
                desktop_path = os.path.join(home_dir, 'Desktop')
            else:  # Linux and others
                desktop_path = os.path.join(home_dir, 'Desktop')
            
            desktop_realpath = os.path.realpath(desktop_path) if os.path.exists(desktop_path) else None
            
            # Allow home directory, Desktop, or subdirectories
            allowed = (resolved_path.startswith(home_realpath + os.sep) or
                      resolved_path == home_realpath)
            
            if desktop_realpath:
                allowed = allowed or (resolved_path.startswith(desktop_realpath + os.sep) or
                                     resolved_path == desktop_realpath)
            
            # Also allow safe paths (temp, working dir)
            return allowed or is_safe_path(path)
        
        elif access_mode == 'unrestricted':
            # Allow any readable directory (dangerous!)
            return os.path.exists(path) and os.access(path, os.R_OK)
        
        return False
        
    except Exception as e:
        logger.warning(f"Error validating path access: {e}")
        return False

# Authentication routes
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page."""
    if request.method == 'POST':
        password = request.form.get('password', '')
        admin_password = get_admin_password()
        
        if password == admin_password:
            session['admin_authenticated'] = True
            flash('Successfully logged in as admin', 'success')
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('monitor_status'))
        else:
            flash('Invalid admin password', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    """Admin logout."""
    session.pop('admin_authenticated', None)
    flash('Successfully logged out', 'success')
    return redirect(url_for('index'))

@app.route('/')
def index():
    # Get recent tracking records
    recent_records = ReflixTracking.query.order_by(ReflixTracking.created_at.desc()).limit(10).all()
    
    # Get summary statistics
    total_references = db.session.query(ReflixTracking.reference_number).distinct().count()
    total_records = ReflixTracking.query.count()
    total_files = LogFile.query.count()
    
    stats = {
        'total_references': total_references,
        'total_records': total_records,
        'total_files': total_files
    }
    
    return render_template('index.html', recent_records=recent_records, stats=stats)

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # Check if file was uploaded
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if file and file.filename and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            
            # Check if file already processed
            existing_file = LogFile.query.filter_by(filename=filename).first()
            if existing_file:
                flash(f'File {filename} has already been processed', 'warning')
                return redirect(url_for('upload_file'))
            
            # Initialize log_file variable
            log_file = None
            
            try:
                # Read file content with progress info
                logger.info(f"Starting to read uploaded file: {filename}")
                file_content = file.read().decode('utf-8')
                file_size = len(file_content)
                logger.info(f"File read complete. Size: {file_size:,} bytes")
                
                # Create log file record
                log_file = LogFile()
                log_file.filename = filename
                log_file.file_size = file_size
                db.session.add(log_file)
                db.session.commit()
                
                # Parse the file
                parser = ReflixLogParser()
                records = parser.parse_log_file(file_content, filename)
                
                # Save tracking records (optimized for large files)
                records_saved = 0
                batch_size = 100
                
                for i, record_data in enumerate(records):
                    try:
                        # Clean and validate data
                        clean_data = {}
                        for key, value in record_data.items():
                            if isinstance(value, str):
                                clean_data[key] = value.encode('utf-8', errors='ignore').decode('utf-8')
                            else:
                                clean_data[key] = value
                        
                        # Simplified duplicate check (just reference number for performance)
                        existing = ReflixTracking.query.filter_by(
                            reference_number=clean_data['reference_number']
                        ).first()
                        
                        if not existing:
                            tracking_record = ReflixTracking(**clean_data)
                            db.session.add(tracking_record)
                            records_saved += 1
                            
                        # Commit in batches to avoid timeout
                        if (i + 1) % batch_size == 0:
                            db.session.commit()
                            logger.info(f"Committed batch: {i+1} records processed")
                            
                    except Exception as e:
                        logger.error(f"Error processing record {i+1}: {e}")
                        continue
                
                # Update log file record
                log_file.processed = True
                log_file.records_extracted = records_saved
                
                db.session.commit()
                
                flash(f'Successfully processed {filename}. Extracted {records_saved} tracking records.', 'success')
                return redirect(url_for('tracking'))
                
            except Exception as e:
                db.session.rollback()
                logger.error(f"Error processing file {filename}: {e}")
                
                # Update log file with error
                try:
                    if 'log_file' in locals() and log_file:
                        log_file.error_message = str(e)
                        db.session.commit()
                except:
                    pass  # If we can't update the error, don't fail completely
                
                flash(f'Error processing file: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Invalid file type. Please upload .txt or .log files only.', 'error')
    
    return render_template('upload.html')

@app.route('/tracking')
def tracking():
    # Get search parameters
    search_ref = request.args.get('reference', '').strip()
    search_status = request.args.get('status', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Build query
    query = ReflixTracking.query
    
    if search_ref:
        query = query.filter(ReflixTracking.reference_number.ilike(f'%{search_ref}%'))
    
    if search_status:
        query = query.filter(ReflixTracking.status.ilike(f'%{search_status}%'))
    
    # Order by timestamp descending
    query = query.order_by(ReflixTracking.timestamp.desc())
    
    # Paginate
    records = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Get unique statuses for filter dropdown
    unique_statuses = db.session.query(ReflixTracking.status).distinct().order_by(ReflixTracking.status).all()
    statuses = [status[0] for status in unique_statuses]
    
    return render_template('tracking.html', 
                         records=records, 
                         statuses=statuses,
                         search_ref=search_ref,
                         search_status=search_status)

@app.route('/reference/<reference_number>')
def reference_detail(reference_number):
    # Get all records for this reference number
    records = ReflixTracking.query.filter_by(
        reference_number=reference_number
    ).order_by(ReflixTracking.timestamp.desc()).all()
    
    if not records:
        flash(f'No records found for reference number {reference_number}', 'warning')
        return redirect(url_for('tracking'))
    
    # Group by shipping unit
    shipping_units = {}
    for record in records:
        unit_key = record.shipping_unit_ref or 'main'
        if unit_key not in shipping_units:
            shipping_units[unit_key] = []
        shipping_units[unit_key].append(record)
    
    return render_template('reference_detail.html', 
                         reference_number=reference_number,
                         shipping_units=shipping_units)

@app.route('/files')
def files():
    # Get all processed files
    files = LogFile.query.order_by(LogFile.upload_timestamp.desc()).all()
    return render_template('files.html', files=files)

@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('500.html'), 500

# Monitor routes
@app.route('/monitor')
@require_admin_auth
def monitor_status():
    """Show monitor status and configuration."""
    # Get monitored folders
    folders = MonitoredFolder.query.all()
    
    # Get monitor instance status
    monitor_instance = MonitorInstance.query.filter_by(id='monitor').first()
    
    # Get file states for each folder
    folder_data = []
    for folder in folders:
        file_states = MonitoredFileState.query.filter_by(folder_id=folder.id).all()
        folder_data.append({
            'folder': folder,
            'file_states': file_states,
            'file_count': len(file_states)
        })
    
    return render_template('monitor_status.html', 
                         folder_data=folder_data,
                         monitor_instance=monitor_instance,
                         is_running=is_monitor_running())

@app.route('/monitor/settings', methods=['GET', 'POST'])
@require_admin_auth
def monitor_settings():
    """Configure monitor settings."""
    if request.method == 'POST':
        folder_path = request.form.get('folder_path', '').strip()
        access_mode = request.form.get('access_mode', 'safe')
        include_patterns = request.form.get('include_patterns', '*.txt,*.log').strip()
        exclude_patterns = request.form.get('exclude_patterns', '').strip()
        polling_interval = int(request.form.get('polling_interval', 10))
        max_files = int(request.form.get('max_files', 10))
        rotation_base = request.form.get('rotation_base', '').strip() or None
        rotation_max = int(request.form.get('rotation_max', 10))
        schedule_enabled = 'schedule_enabled' in request.form
        schedule_every_minutes = int(request.form.get('schedule_every_minutes', 120))
        active = 'active' in request.form
        
        if not folder_path:
            flash('Folder path is required', 'error')
            return redirect(url_for('monitor_settings'))
        
        # Validate path based on access mode
        if not validate_path_access(folder_path, access_mode):
            if access_mode == 'safe':
                flash('Folder path is not allowed for security reasons. Only paths under /tmp or current working directory are permitted.', 'error')
            elif access_mode == 'home_desktop':
                flash('Desktop folder access is restricted to your home Desktop directory only.', 'error')
            else:
                flash('Invalid folder path or access denied.', 'error')
            return redirect(url_for('monitor_settings'))
        
        if not os.path.exists(folder_path):
            flash(f'Folder path does not exist: {folder_path}', 'error')
            return redirect(url_for('monitor_settings'))
        
        try:
            # Check if folder already exists
            existing_folder = MonitoredFolder.query.filter_by(path=folder_path).first()
            
            if existing_folder:
                # Update existing folder
                existing_folder.access_mode = access_mode
                existing_folder.include_patterns = include_patterns
                existing_folder.exclude_patterns = exclude_patterns or None
                existing_folder.polling_interval = polling_interval
                existing_folder.max_files = max_files
                existing_folder.rotation_base = rotation_base
                existing_folder.rotation_max = rotation_max
                existing_folder.schedule_enabled = schedule_enabled
                existing_folder.schedule_every_minutes = schedule_every_minutes
                existing_folder.active = active
                flash(f'Updated folder monitoring settings for {folder_path}', 'success')
            else:
                # Create new folder
                folder = MonitoredFolder()
                folder.path = folder_path
                folder.access_mode = access_mode
                folder.include_patterns = include_patterns
                folder.exclude_patterns = exclude_patterns or None
                folder.polling_interval = polling_interval
                folder.max_files = max_files
                folder.rotation_base = rotation_base
                folder.rotation_max = rotation_max
                folder.schedule_enabled = schedule_enabled
                folder.schedule_every_minutes = schedule_every_minutes
                folder.active = active
                db.session.add(folder)
                flash(f'Added folder for monitoring: {folder_path}', 'success')
            
            db.session.commit()
            
            # Start monitor if it's not running and we have active folders
            if active and not is_monitor_running():
                if start_monitor():
                    flash('Monitor service started', 'success')
                else:
                    flash('Failed to start monitor service', 'warning')
            
        except Exception as e:
            logger.error(f"Error updating monitor settings: {e}")
            flash(f'Error updating settings: {str(e)}', 'error')
            db.session.rollback()
        
        return redirect(url_for('monitor_status'))
    
    # GET request - show settings form
    folders = MonitoredFolder.query.all()
    return render_template('monitor_settings.html', folders=folders)


@app.route('/analytics')
def analytics():
    """Analytics dashboard with timestamp-based charts and statistics."""
    from sqlalchemy import func, text
    from datetime import datetime, timedelta
    
    # Get date range from query parameters (default: last 30 days)
    end_date = request.args.get('end_date')
    start_date = request.args.get('start_date')
    
    if not end_date:
        end_date = datetime.now().strftime('%Y-%m-%d')
    if not start_date:
        start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    # Convert to datetime objects for queries
    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    
    # Get basic statistics
    total_records = ReflixTracking.query.filter(
        ReflixTracking.created_at.between(start_datetime, end_datetime)
    ).count()
    
    unique_references = ReflixTracking.query.filter(
        ReflixTracking.created_at.between(start_datetime, end_datetime)
    ).with_entities(ReflixTracking.reference_number).distinct().count()
    
    # Status breakdown
    status_data = db.session.query(
        ReflixTracking.status,
        func.count(ReflixTracking.id).label('count')
    ).filter(
        ReflixTracking.created_at.between(start_datetime, end_datetime)
    ).group_by(ReflixTracking.status).all()
    
    # Daily tracking volume (last 30 days)
    daily_volume = db.session.query(
        func.date(ReflixTracking.created_at).label('date'),
        func.count(ReflixTracking.id).label('count')
    ).filter(
        ReflixTracking.created_at.between(start_datetime, end_datetime)
    ).group_by(func.date(ReflixTracking.created_at)).order_by('date').all()
    
    # Top reference numbers by activity
    top_references = db.session.query(
        ReflixTracking.reference_number,
        func.count(ReflixTracking.id).label('count'),
        func.max(ReflixTracking.created_at).label('last_update')
    ).filter(
        ReflixTracking.created_at.between(start_datetime, end_datetime)
    ).group_by(ReflixTracking.reference_number).order_by(text('count DESC')).limit(10).all()
    
    # Recent activity
    recent_activity = ReflixTracking.query.filter(
        ReflixTracking.created_at.between(start_datetime, end_datetime)
    ).order_by(ReflixTracking.created_at.desc()).limit(20).all()
    
    # Calculate days in range
    days_in_range = (end_datetime.date() - start_datetime.date()).days
    
    return render_template('analytics.html',
                         total_records=total_records,
                         unique_references=unique_references,
                         status_data=status_data,
                         daily_volume=daily_volume,
                         top_references=top_references,
                         recent_activity=recent_activity,
                         start_date=start_date,
                         end_date=end_date,
                         days_in_range=days_in_range)


@app.route('/export.xlsx')
def export_excel():
    """Export tracking data to Excel with filtering and multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime, timedelta
    import io
    
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status_filter = request.args.get('status')
    reference_filter = request.args.get('reference')
    
    # Default to last 30 days if no dates provided
    if not end_date:
        end_date = datetime.now().strftime('%Y-%m-%d')
    if not start_date:
        start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    # Convert to datetime objects
    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    
    # Build query with filters
    query = ReflixTracking.query.filter(
        ReflixTracking.created_at.between(start_datetime, end_datetime)
    )
    
    if status_filter:
        query = query.filter(ReflixTracking.status == status_filter)
    if reference_filter:
        query = query.filter(ReflixTracking.reference_number.contains(reference_filter))
    
    records = query.order_by(ReflixTracking.created_at.desc()).all()
    
    # Create workbook with multiple sheets
    wb = Workbook()
    
    # Sheet 1: Records Data
    ws_records = wb.active
    ws_records.title = "Tracking Records"
    
    # Headers with styling
    headers = [
        'Reference Number', 'Shipping Unit Ref', 'Status', 'Description',
        'Location', 'Timestamp', 'Created At', 'Log File'
    ]
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_records.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data rows
    for row_num, record in enumerate(records, 2):
        ws_records.cell(row=row_num, column=1, value=record.reference_number)
        ws_records.cell(row=row_num, column=2, value=record.shipping_unit_ref or '')
        ws_records.cell(row=row_num, column=3, value=record.status)
        ws_records.cell(row=row_num, column=4, value=record.description or '')
        ws_records.cell(row=row_num, column=5, value=record.location or '')
        ws_records.cell(row=row_num, column=6, value=record.timestamp)
        ws_records.cell(row=row_num, column=7, value=record.created_at)
        ws_records.cell(row=row_num, column=8, value=record.log_file_name or '')
    
    # Auto-adjust column widths
    for column in ws_records.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_records.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Summary Statistics
    ws_summary = wb.create_sheet(title="Summary")
    
    # Summary data
    from sqlalchemy import func
    
    total_records = len(records)
    unique_references = len(set(record.reference_number for record in records))
    
    # Status counts
    status_counts = db.session.query(
        ReflixTracking.status,
        func.count(ReflixTracking.id).label('count')
    ).filter(
        ReflixTracking.created_at.between(start_datetime, end_datetime)
    ).group_by(ReflixTracking.status).all()
    
    # Daily counts
    daily_counts = db.session.query(
        func.date(ReflixTracking.created_at).label('date'),
        func.count(ReflixTracking.id).label('count')
    ).filter(
        ReflixTracking.created_at.between(start_datetime, end_datetime)
    ).group_by(func.date(ReflixTracking.created_at)).order_by('date').all()
    
    # Summary headers
    ws_summary.cell(row=1, column=1, value='REFLIV Tracking Export Summary')
    ws_summary.cell(row=1, column=1).font = Font(size=16, bold=True)
    
    ws_summary.cell(row=3, column=1, value=f'Export Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    ws_summary.cell(row=4, column=1, value=f'Date Range: {start_date} to {end_date}')
    ws_summary.cell(row=5, column=1, value=f'Total Records: {total_records}')
    ws_summary.cell(row=6, column=1, value=f'Unique References: {unique_references}')
    
    # Status breakdown
    ws_summary.cell(row=8, column=1, value='Status Breakdown:')
    ws_summary.cell(row=8, column=1).font = Font(bold=True)
    
    for i, (status, count) in enumerate(status_counts, 9):
        ws_summary.cell(row=i, column=1, value=f'{status}: {count}')
    
    # Daily breakdown
    start_row = len(status_counts) + 11
    ws_summary.cell(row=start_row, column=1, value='Daily Breakdown:')
    ws_summary.cell(row=start_row, column=1).font = Font(bold=True)
    
    ws_summary.cell(row=start_row + 1, column=1, value='Date')
    ws_summary.cell(row=start_row + 1, column=2, value='Records')
    
    for cell in ws_summary[f'A{start_row + 1}:B{start_row + 1}'][0]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    
    for i, (date, count) in enumerate(daily_counts, start_row + 2):
        ws_summary.cell(row=i, column=1, value=str(date))
        ws_summary.cell(row=i, column=2, value=count)
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create filename with timestamp and filters
    filename_parts = ['refliv_tracking']
    filename_parts.append(f'{start_date}_to_{end_date}')
    if status_filter:
        filename_parts.append(f'status_{status_filter}')
    if reference_filter:
        filename_parts.append(f'ref_{reference_filter[:10]}')
    filename = '_'.join(filename_parts) + '.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/monitor/start', methods=['POST'])
@require_admin_auth
def start_monitor_service():
    """Start the folder monitor service."""
    try:
        if start_monitor():
            flash('Monitor service started successfully', 'success')
        else:
            flash('Monitor service is already running or failed to start', 'warning')
    except Exception as e:
        flash(f'Error starting monitor: {str(e)}', 'error')
    
    return redirect(url_for('monitor_status'))

@app.route('/monitor/stop', methods=['POST'])
@require_admin_auth
def stop_monitor_service():
    """Stop the folder monitor service."""
    try:
        stop_monitor()
        flash('Monitor service stopped', 'success')
    except Exception as e:
        flash(f'Error stopping monitor: {str(e)}', 'error')
    
    return redirect(url_for('monitor_status'))

@app.route('/monitor/folders/<int:folder_id>/delete', methods=['POST'])
@require_admin_auth
def delete_monitored_folder(folder_id):
    """Delete a monitored folder and its file states."""
    try:
        folder = MonitoredFolder.query.get_or_404(folder_id)
        folder_path = folder.path
        
        # Delete folder (cascades to file states)
        db.session.delete(folder)
        db.session.commit()
        
        flash(f'Removed folder from monitoring: {folder_path}', 'success')
        
    except Exception as e:
        logger.error(f"Error deleting monitored folder: {e}")
        flash(f'Error removing folder: {str(e)}', 'error')
        db.session.rollback()
    
    return redirect(url_for('monitor_status'))
